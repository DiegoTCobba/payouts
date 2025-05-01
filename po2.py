import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
import io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import re

st.title("RECHAZOS DE PAGOS MASIVOS")

# Crear pestañas
preRech, postRech = st.tabs(["📄 PRE RECHAZOS BCP", "📄 POST RECHAZOS"])


with preRech:
    st.subheader("PRE RECHAZOS BCP")
    st.divider()
    st.write("Herramienta diseñada para extraer clientes desde un archivo PDF y filtrar automáticamente en la base de datos Excel")

    # Subir de archivos pdf y excel
    st.markdown("### ARCHIVO PDF")
    pdf_file = st.file_uploader("Sube un archivo PDF", type="pdf")
    st.markdown("### MASIVO EXCEL")
    excel_file = st.file_uploader("Sube un archivo Excel", type=["xlsx"])

    if pdf_file and excel_file:
        # Extraer texto del PDF
        text = ""
        with fitz.open(stream=pdf_file.read(), filetype="pdf") as doc:
            for page in doc:
                text += page.get_text()

        # Eliminar posibles números de cuenta como 380-75148297-0-31
        cuentas_posibles = re.findall(r'\b[\d]{2,4}-[\d]{5,10}-\d{1,2}-\d{1,3}\b', text)
        for cuenta in cuentas_posibles:
            text = text.replace(cuenta, '')

        # Buscar números de documento de 6 o más dígitos
        numeros_documento = re.findall(r'\b\d{6,}\b', text)
        numeros_documento = list(set(numeros_documento))  # eliminar duplicados
        documentos_set = set(numeros_documento)

        # Reposicionar puntero para openpyxl
        excel_file.seek(0)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Resaltado relleno amarillo
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Columnas a ocultar en el excel
        columnas_a_ocultar = ['B', 'C', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'N', 'O', 'P','Q']
        for col in columnas_a_ocultar:
            ws.column_dimensions[col].hidden = True

        # Filtrar filas con coincidencias
        filas_con_coincidencias = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if any(str(cell.value) in documentos_set for cell in row):
                filas_con_coincidencias.append([cell.value for cell in row])

        # Crear excel nuevo
        wb_filtrado = Workbook()
        ws_filtrado = wb_filtrado.active

        # Encabezados
        encabezados = [cell.value for cell in ws[1]]
        ws_filtrado.append(encabezados)

        # Agregar filas filtradas
        for row in filas_con_coincidencias:
            ws_filtrado.append(row)

        # Aplicar resaltado
        for row in ws_filtrado.iter_rows(min_row=2):
            for cell in row:
                if str(cell.value) in documentos_set:
                    cell.fill = fill

        # Ocultar columnas
        for col in columnas_a_ocultar:
            ws_filtrado.column_dimensions[col].hidden = True

        #Mantener ancho de columnas no ocultas
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            letra = get_column_letter(col_idx)
            if letra not in columnas_a_ocultar:
                if letra in ws.column_dimensions and ws.column_dimensions[letra].width:
                    ancho = ws.column_dimensions[letra].width
                    ws_filtrado.column_dimensions[letra].width = ancho

        # Guardar nuevo archivo en memoria
        output_filtrado = io.BytesIO()
        wb_filtrado.save(output_filtrado)
        output_filtrado.seek(0)

        # Mostrar vista previa
        df_filtrado = pd.DataFrame(filas_con_coincidencias, columns=encabezados)
        letras_a_indices = [ord(c) - ord('A') for c in columnas_a_ocultar]
        columnas_visibles = [col for idx, col in enumerate(df_filtrado.columns) if idx not in letras_a_indices]
        df_visible = df_filtrado[columnas_visibles]

        st.subheader("Vista previa final con filas y columnas filtradas:")
        st.dataframe(df_visible)

        # Botón descarga
        st.download_button(
            label="Descargar Excel con resaltado y filas filtradas",
            data=output_filtrado,
            file_name="resaltado_filtrado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#--------------------------------------------------------------------------------------
with postRech:
    st.subheader("POST RECHAZOS")
    st.divider()
    st.write("Herramienta diseñada para extraer DNI desde un archivo PDF y filtrar automáticamente en la base de datos de un Excel")

    # Subir de archivos pdf y excel
    st.markdown("### ARCHIVO PDF")
    pdf_file = st.file_uploader("Sube un archivo PDF", type="pdf")
    st.markdown("### MASIVO EXCEL")
    excel_file = st.file_uploader("Sube un archivo Excel", type=["xlsx"])

    if pdf_file and excel_file:
        # Extraer texto del PDF
        text = ""
        with fitz.open(stream=pdf_file.read(), filetype="pdf") as doc:
            for page in doc:
                text += page.get_text()

        # Eliminar posibles números de cuenta como 380-75148297-0-31
        cuentas_posibles = re.findall(r'\b[\d]{2,4}-[\d]{5,10}-\d{1,2}-\d{1,3}\b', text)
        for cuenta in cuentas_posibles:
            text = text.replace(cuenta, '')

        # Buscar números de documento de 6 o más dígitos
        numeros_documento = re.findall(r'\b\d{6,}\b', text)
        numeros_documento = list(set(numeros_documento))  # eliminar duplicados
        documentos_set = set(numeros_documento)

        # Reposicionar puntero para openpyxl
        excel_file.seek(0)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Resaltado relleno amarillo
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Columnas a ocultar en el excel
        columnas_a_ocultar = ['B', 'C', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'N', 'O', 'P','Q']
        for col in columnas_a_ocultar:
            ws.column_dimensions[col].hidden = True

        # Filtrar filas con coincidencias
        filas_con_coincidencias = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if any(str(cell.value) in documentos_set for cell in row):
                filas_con_coincidencias.append([cell.value for cell in row])

        # Crear excel nuevo
        wb_filtrado = Workbook()
        ws_filtrado = wb_filtrado.active

        # Encabezados
        encabezados = [cell.value for cell in ws[1]]
        ws_filtrado.append(encabezados)

        # Agregar filas filtradas
        for row in filas_con_coincidencias:
            ws_filtrado.append(row)

        # Aplicar resaltado
        for row in ws_filtrado.iter_rows(min_row=2):
            for cell in row:
                if str(cell.value) in documentos_set:
                    cell.fill = fill

        # Ocultar columnas
        for col in columnas_a_ocultar:
            ws_filtrado.column_dimensions[col].hidden = True

        #Mantener ancho de columnas no ocultas
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            letra = get_column_letter(col_idx)
            if letra not in columnas_a_ocultar:
                if letra in ws.column_dimensions and ws.column_dimensions[letra].width:
                    ancho = ws.column_dimensions[letra].width
                    ws_filtrado.column_dimensions[letra].width = ancho

        # Guardar nuevo archivo en memoria
        output_filtrado = io.BytesIO()
        wb_filtrado.save(output_filtrado)
        output_filtrado.seek(0)

        # Mostrar vista previa
        df_filtrado = pd.DataFrame(filas_con_coincidencias, columns=encabezados)
        letras_a_indices = [ord(c) - ord('A') for c in columnas_a_ocultar]
        columnas_visibles = [col for idx, col in enumerate(df_filtrado.columns) if idx not in letras_a_indices]
        df_visible = df_filtrado[columnas_visibles]

        st.subheader("Vista previa final con filas y columnas filtradas:")
        st.dataframe(df_visible)

        # Botón descarga
        st.download_button(
            label="Descargar Excel con resaltado y filas filtradas",
            data=output_filtrado,
            file_name="resaltado_filtrado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    #--------------------------------------------------------------------------------------
