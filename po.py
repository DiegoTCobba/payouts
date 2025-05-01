import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

st.title("📄🔍 Filtrado y resaltado de clientes a rechazar")

# Subir archivos
pdf_file = st.file_uploader("Sube un archivo PDF", type="pdf")
excel_file = st.file_uploader("Sube un archivo Excel", type=["xlsx"])

if pdf_file and excel_file:
    # Extraer texto del PDF
    text = ""
    with fitz.open(stream=pdf_file.read(), filetype="pdf") as doc:
        for page in doc:
            text += page.get_text()

    # Eliminar posibles números de cuenta como 380-75148297-0-31
    cuentas_posibles = re.findall(r'\b[\d]{2,4}-[\d]{5,10}-\d{1,2}-\d{1,3}\b', text)
    for cuenta in cuentas_posibles:
        text = text.replace(cuenta, '')

    # Buscar números de documento puros de 6 o más dígitos (sin símbolos ni guiones)
    numeros_documento = re.findall(r'\b\d{6,}\b', text)
    numeros_documento = list(set(numeros_documento))  # eliminar duplicados

    # Leer Excel con pandas (previa visualización)
    df = pd.read_excel(excel_file)

    # Reposicionar puntero para openpyxl
    output = io.BytesIO()
    excel_file.seek(0)
    wb = load_workbook(excel_file)
    ws = wb.active

    # Estilo de resaltado amarillo
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Ocultar columnas específicas en el archivo Excel
    columnas_a_ocultar = ['B', 'C', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'N', 'O', 'P', 'R']
    for col in columnas_a_ocultar:
        ws.column_dimensions[col].hidden = True
    
    # Resaltar coincidencias
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value) in numeros_documento:
                cell.fill = fill

    # Eliminar las filas donde la columna A tiene celdas sin relleno amarillo
    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Saltar encabezado
        if ws[f"A{i}"].fill != fill:  # Si la celda en la columna A no tiene relleno amarillo
            rows_to_delete.append(i)

    # Eliminar las filas desde el final hacia el principio (para no afectar los índices al eliminar)
    for row_index in reversed(rows_to_delete):
        ws.delete_rows(row_index)

    # Filtrar filas con al menos una celda resaltada en amarillo
    filas_resaltadas = []
    for row in ws.iter_rows(min_row=2):  # Saltamos el encabezado
        if ws[f"A{row[0].row}"].fill == fill:  # Verificamos si la columna A tiene relleno amarillo
            filas_resaltadas.append([cell.value for cell in row])

    # Convertir filas filtradas a DataFrame
    columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    df_resaltado = pd.DataFrame(filas_resaltadas, columns=columns)

    # Ocultar columnas específicas por letra
    columnas_a_ocultar = ['B', 'C', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'N', 'O', 'P', 'R']
    
    # Convertir letras a índices (0-based)
    letras_a_indices = [ord(c) - ord('A') for c in columnas_a_ocultar]

    # Eliminar del DataFrame
    columnas_visibles = [col for idx, col in enumerate(df_resaltado.columns) if idx not in letras_a_indices]
    df_visible = df_resaltado[columnas_visibles]

    # Mostrar DataFrame filtrado en la app
    st.subheader("📊 Vista previa final con columnas ocultas:")
    st.dataframe(df_visible)

    # Guardar archivo en memoria (ahora sin filas blancas)
    wb.save(output)
    output.seek(0)

    # Botón de descarga
    st.download_button(
        label="📥 Descargar Excel con resaltado",
        data=output,
        file_name="resaltado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


